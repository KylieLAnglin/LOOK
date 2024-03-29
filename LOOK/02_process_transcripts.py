# %%
import os
import re

import pandas as pd
import numpy as np
from openpyxl import load_workbook

from library import start

# %%
df = pd.read_csv(start.DATA_DIR + "clean/meta_data.csv")
# df = df.head(11)
# %%


def generate_transcript_df(string, filename):
    list_of_utterances = string.split("\n")

    list_of_utterances = [item for item in list_of_utterances if item != ""]

    transcript_df = pd.DataFrame()

    transcript_df["original_text"] = list_of_utterances

    transcript_df["without_brackets"] = transcript_df["original_text"].apply(
        lambda x: re.sub("\[.*?\]", "", x)
    )

    # transcript_df["without_brackets"] = transcript_df["original_text"].apply(
    #     lambda x: re.sub("\[[^:]*\]", "", x)
    # )

    transcript_df["speaker"] = transcript_df["without_brackets"].str.split(":").str[0]
    transcript_df["text"] = transcript_df["without_brackets"].str.split(":").str[1]

    transcript_df["text"] = np.where(
        transcript_df.text.isnull(), transcript_df.original_text, transcript_df.text
    )

    transcript_df["speaker_coach"] = np.where(
        transcript_df.speaker.str.contains("Coach"), 1, 0
    )

    transcript_df["coach_count"] = transcript_df["speaker_coach"].cumsum()

    return transcript_df


# %%
def transcript_df_to_excel(transcript_df, excel_template: str, excel_file: str):

    wb = load_workbook(excel_template)
    ws = wb.active

    row = 2
    col = 1
    for speaker in transcript_df.speaker:
        ws.cell(row=row, column=col).value = str(speaker)
        row = row + 1

    row = 2
    col = 2

    for count in transcript_df.coach_count:
        ws.cell(row=row, column=col).value = count
        row = row + 1

    row = 2
    col = 3

    for text in transcript_df.without_brackets:
        ws.cell(row=row, column=col).value = text
        row = row + 1

    wb.save(excel_file)


# %%
# test_file = list(df.file)[0]

# with open(
#     start.DATA_DIR + "raw/transcripts/" + test_file + ".txt",
#     encoding="ISO-8859-1",
# ) as file:
#     text_str = file.read()


# test = generate_transcript_df(text_str, filename=test_file)

# transcript_df_to_excel(
#     test,
#     excel_template=start.DATA_DIR + "raw/transcript_coding_template.xlsx",
#     excel_file=start.DATA_DIR + "clean/transcripts_to_excel/" + test_file + ".xlsx",
# )

# %%
for filename in df.file:

    with open(
        start.DATA_DIR + "raw/transcripts/" + filename + ".txt",
        encoding="ISO-8859-1",
    ) as file:
        text_str = file.read()

    transcript_df = generate_transcript_df(text_str, filename=filename)
    print(filename)
    # if test.speaker.nunique() > 5:
    #     print(test.speaker.value_counts())
    transcript_df_to_excel(
        transcript_df,
        excel_template=start.DATA_DIR + "raw/transcript_coding_template.xlsx",
        excel_file=start.DATA_DIR + "clean/transcripts_to_excel/" + filename + ".xlsx",
    )
transcript_df.speaker.value_counts()
# %%
transcript_df
# %%
