# %%
import os
import re
import fnmatch
import collections

import pandas as pd
import docx
from openpyxl import load_workbook

import random
from moves.library import start
from moves.library import process_transcripts

# %%

# %%
def transcript_df_to_excel(transcript_df, excel_template: str, excel_file: str):

    wb = load_workbook(excel_template)
    ws = wb.active

    row = 2
    col = 1
    for time in transcript_df.time:
        ws.cell(row=row, column=col).value = str(time)
        row = row + 1

    row = 2
    col = 2

    for speaker in transcript_df.speaker:
        ws.cell(row=row, column=col).value = str(speaker)
        row = row + 1

    row = 2
    col = 3
    for text in transcript_df.text:
        ws.cell(row=row, column=col).value = text
        row = row + 1

    wb.save(excel_file)


# %%
speaker_tags = pd.read_excel(start.DATA_PATH + "speaker_tags.xlsx")

speaker_tags = {
    doc: coach for (doc, coach) in zip(speaker_tags.doc, speaker_tags.coach)
}
speaker_tags = {doc: speaker[:-1] for (doc, speaker) in speaker_tags.items()}

# %%

# PATH = "/Users/kylie/Dropbox/1 Projects/Coaching Moves/"
# doc_file = "29_c_Transcript-updated"

# transcript_to_excel(
#     excel_template=PATH + "template.xlsx",
#     doc_file=PATH + doc_file + ".docx",
#     excel_file=PATH + doc_file + ".xlsx",
# )

files = [
    filename
    for filename in os.listdir(start.CC_PATH + "uncoded transcripts/")
    if fnmatch.fnmatch(filename, "*.docx") and not filename.startswith("~$")
]

for filename in files:
    # paragraphs = extract_paragraphs(doc_file=start.TRANSCRIPTS_PATH + filename)
    # transcript = extract_data_from_go_transcript(turns_of_talk=paragraphs)
    transcript = process_transcripts.word_to_transcript(
        doc_file=start.TRANSCRIPTS_PATH + filename
    )
    if transcript:
        transcript_df = process_transcripts.transcript_to_cleaned_df(
            transcript=transcript
        )
        speakers = [
            speaker.replace(speaker_tags[filename], "Coach")
            for speaker in transcript_df.speaker
        ]
        transcript_df["speaker"] = speakers
        filename = filename[:-5]
        transcript_df_to_excel(
            transcript_df=transcript_df,
            excel_template=start.CC_PATH + "template.xlsx",
            excel_file=start.CC_PATH + "excel transcripts/" + filename + ".xlsx",
        )


# %%

# Create random order of files for coding
random.seed(10)
random.shuffle(files)
random_df = pd.DataFrame(files)
random_df.to_csv(start.CC_PATH + "random_order.csv")
